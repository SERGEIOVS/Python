from fileinput import filename
from typing import List
from openpyxl import load_workbook



Result = [



'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',

'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',

'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy',
'ttt' , 'xx' , 'yy'



]



start = True
while start:

    
            # Load in the workbook
            filename = input('nazvanie faila : ')

            wb = load_workbook(str(filename))
            wb.active = 1
            sheet = wb.active
            
            filemode = 'w'
            f = open(str(filename),filemode)

            # Get sheet names
            for i in range(1,len(Result)):
                sheet['A'+str(i)].value =str(Result[i])
                print(sheet['A'+str(i)].value)

            newfilename = input('kuda sohranjaem fail? : ')
            wb.save(str(newfilename))





"""
from typing import List
from openpyxl import load_workbook

List1 = [ 'ttt' , 'b' , 'a' ]

# Load in the workbook
wb = load_workbook('123.xlsx')
wb.active = 1

sheet = wb.active

# Get sheet names

for i in range(1,len(List1)):
    sheet['A'+str(i)].value =str(List1[i])
    print(sheet['A'+str(i)].value)
    
wb.save('123.xlsx')



#Создаем листы/sheets в файле exel

from openpyxl import load_workbook
# Load in the workbook
wb = load_workbook('расчет.xlsx')
data = 'My'
# Get sheet names
wb.create_sheet(data)
print(wb.get_sheet_names())





import statistics 
import pandas as pd

my_file = open('кол-во детонаторов.txt', 'w')

det_number   = [1 , 2 , 3 , 4 , 5 , 6 , 7 , 8 , 9 , 10 , 11 , 12 , 13 , 14 , 15 , 16 , 17 , 18]

det_quantity = [1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 , 1 ,  1 ,  1 ,  1 ,  1 ,  1 ,  2 ,  1 ,  1]

camera_V = 60

print()

print()

run = True

while run:
    print()
    Total_need =int(input('Cколько кубов требуется добыть? '))

    #12640 кубов

    Camers_count = Total_need / camera_V

    print('',int(Camers_count),'камер нужно')

    print()

    print('|','-' * 50)

    for i in range(len(det_quantity)):

        print('|','детонаторов серии',det_number[i],'нужно:',det_quantity[i] *int(Camers_count),'|')

        print('-' * 50)

        vsego = [det_number[i],det_quantity[i] * int(camera_V)]
    
        my_file.write(str(vsego)+'\n')

        df = pd.DataFrame({

                            'Тип'    : det_number,

                            'Кол-во' :  det_quantity

                          })
        
    df.to_excel('./расчет.xlsx')

    my_file.close()
    print( df['Кол-во'] )
    
    print(df)




"""
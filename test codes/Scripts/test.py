from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('расчет.xlsx')
data = 'My'
# Get sheet names
wb.create_sheet(data)
print(wb.get_sheet_names())


"""

import openpyxl
wb = openpyxl.reader.excel.load_workbook(filename="sample.xlsx", data_only=True)
print(wb.sheetnames)
wb.active = 1
sheet = wb.active
#print(sheet['A1'].value)
for i in range(1,12):
    print(sheet['A'+str(i)].value,sheet['B'+str(i)].value,sheet['C'+str(i)].value)

"""
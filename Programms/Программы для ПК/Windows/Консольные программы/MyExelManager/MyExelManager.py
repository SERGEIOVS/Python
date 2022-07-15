from openpyxl import load_workbook

filemodes = ['r','w']

currentfm = 0

letters = ['A']

nums = [1,2,3,4,5,6,7,8,9]

currentnum = 0
letter = letters[0]

#main programm cycle
start = True
while start:
            #READ FILE

            action = input('action : ')

            row = int(input('row : '))

            letter = input('letter : ')
    

            

            if action =='read file':

                filename = input('nazvanie faila : ')


                # Load in the workbook
                #wb -> workbook
                wb = load_workbook(str(filename))

                filemode = filemodes[currentfm]
                
                file = open(str(filename),filemode)
                
                #wb -> workbook
                #ws -> worksheet

                ws = wb.active
                                
                print(wb.sheetnames)

                print(ws[letter + str(nums[currentnum])].value)

                print(file.name)

                file.close()




            #WRITE FILE
            if action =='write file':

                filemode = filemodes[currentfm]

                filename = input('nazvanie faila : ')

                myvalue = input('value : ')

                f = open(str(filename),filemode)

                wb = load_workbook(str(filename))

                ws = wb.active
                    
                d = ws.cell(row = letter, column = row, value = myvalue)

                print(d.value)

                f.close()
                










"""            #set a file mode 'w'- write in file  , 'r'- read file 
            if action =='write file':
                filemode = 'w'
                f = open(str(filename),filemode)

                # Get sheet names
                #writing values from results in exel file

                for i in range(1,len(Results)):
                    sheet['A'+str(i)].value =str(Results[i])
                    print(sheet['A' + str(i)].value)
                f.close()

"""
                